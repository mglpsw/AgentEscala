from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from functools import lru_cache
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..models import MedicalProfile, Shift, User

_MONTHS_PT = {
    "janeiro": 1,
    "fevereiro": 2,
    "março": 3,
    "marco": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

_COL_SYNONYMS: Dict[str, List[str]] = {
    "professional": ["profissional", "médico", "medico", "nome", "plantonista", "nome completo"],
    "crm": ["crm", "conselho", "registro"],
    "date": ["data", "dia", "plantão", "plantao"],
    "start_time": ["entrada", "início", "inicio", "hora início", "hora inicio"],
    "end_time": ["saída", "saida", "fim", "hora fim", "hora saída", "hora saida"],
    "shift_label": ["turno", "período", "periodo", "plantão", "plantao"],
    "specialty": ["especialidade"],
    "unit": ["unidade", "local", "setor"],
    "duration": ["total horas", "carga horária", "carga horaria", "horas"],
    "weekday": ["dia semana", "dia da semana"],
}

NAME_ALIASES = {
    "cloves domingos": "Cloves Domingos Rufino",
    "mariana koppe": "Mariana Koppe Pereira",
    "jean pierri": "Jean Pierre Dosciatti",
    "leticia leonarda g.cravo": "Leticia Leonarda",
    "joel dahne": "Joel Soares Dahne",
    "leticia leonarda": "Leticia Leonarda",
    "leticia": "Leticia Leonarda",
    "jean": "Jean Pierre Dosciatti",
}

SHIFT_KIND_COLORS = {
    "day": "green",
    "intermediate": "yellow",
    "night": "blue",
    "twenty_four": "purple",
    "custom": "gray",
}

_PHONE_RE = re.compile(r"(?:\+?55\s*)?(?:\(?\d{2}\)?\s*)?\d{4,5}[-\s]?\d{4}")
_CRM_RE = re.compile(r"\bCRM\s*[:\-/]?\s*([A-Z]{0,2})\s*(\d{3,8})\b", re.IGNORECASE)
_CRM_NUMERIC_RE = re.compile(r"\b(\d{4,8})(?:\s*e\s*(\d{3,8}))?\b", re.IGNORECASE)
_DATE_RE = re.compile(r"^(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?$")
_TIME_RE = re.compile(r"(\d{1,2})(?::|h)?(\d{2})?")
_OPERATIONAL_SUFFIX_RE = re.compile(
    r"\b(?:faturamento|apoio|coringa|extra|plantao extra|disponivel|cobertura|troca|folga)\b.*$",
    re.IGNORECASE,
)


_NORMALIZATION_CACHE: dict[str, str] = {}
_MATCH_SCORE_CACHE: dict[tuple[str, str], float] = {}


def _normalize_text(value: Optional[str]) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def _comparable_text(value: Optional[str]) -> str:
    key = _normalize_text(value).lower()
    if key in _NORMALIZATION_CACHE:
        return _NORMALIZATION_CACHE[key]
    normalized = "".join(c for c in unicodedata.normalize("NFKD", key) if not unicodedata.combining(c))
    _NORMALIZATION_CACHE[key] = normalized
    return normalized


@lru_cache(maxsize=1024)
def _canonicalize_name(value: str) -> str:
    cmp = _comparable_text(value)
    return " ".join(part.capitalize() for part in cmp.split())


def _extract_month_year(label: str) -> Tuple[Optional[int], Optional[int], float]:
    label_cmp = _comparable_text(label)
    year_match = re.search(r"\b(20\d{2})\b", label_cmp)
    year = int(year_match.group(1)) if year_match else None
    for name, month in _MONTHS_PT.items():
        if name in label_cmp:
            return month, year, 0.9 if year else 0.7
    slash = re.search(r"\b(\d{1,2})\s*/\s*(20\d{2})\b", label_cmp)
    if slash:
        return int(slash.group(1)), int(slash.group(2)), 0.85
    return None, year, 0.0


def _canonical_header(header: str) -> Optional[str]:
    cmp = _comparable_text(header)
    for canonical, synonyms in _COL_SYNONYMS.items():
        for s in synonyms:
            if _comparable_text(s) in cmp:
                return canonical
    return None


def _split_compound_values(value: Optional[str]) -> List[str]:
    clean = _normalize_text(value)
    if not clean:
        return []
    pieces = re.split(r"\s+e\s+|\s*/\s*|\s*\+\s*|\s*&\s*|\s*;\s*|\s*,\s*", clean, flags=re.IGNORECASE)
    return [_normalize_text(piece) for piece in pieces if _normalize_text(piece)]


def _clean_professional(value: Optional[str]) -> Tuple[str, Optional[str], List[str], str, bool]:
    raw = _normalize_text(value)
    messages: List[str] = []
    if not raw:
        return "", None, ["Nome do profissional ausente"], "", False

    crm = None
    crm_match = _CRM_RE.search(raw)
    if crm_match:
        uf = (crm_match.group(1) or "").upper()
        num = crm_match.group(2)
        crm = f"{uf}{num}" if uf else num
        raw = _CRM_RE.sub("", raw)

    cleaned = _PHONE_RE.sub("", raw)
    cleaned = _OPERATIONAL_SUFFIX_RE.sub("", cleaned)
    cleaned = re.sub(r"\(.*?\)", "", cleaned)
    cleaned = re.sub(r"[^\wÀ-ÿ.\-\s]", " ", cleaned)
    cleaned = _normalize_text(cleaned)

    canonical = _canonicalize_name(cleaned)
    alias_applied = False
    alias = NAME_ALIASES.get(_comparable_text(cleaned))
    if alias:
        canonical = alias
        alias_applied = True
        messages.append("Alias aplicado")

    if len(cleaned.split()) < 2:
        messages.append("Nome com baixa qualidade")

    return cleaned, crm, messages, canonical, alias_applied


def _parse_date_with_context(raw_date: Optional[str], month: Optional[int], year: Optional[int]) -> Optional[date]:
    if raw_date:
        raw_date = _normalize_text(raw_date)
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M"):
            try:
                return datetime.strptime(raw_date, fmt).date()
            except ValueError:
                pass
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                parsed = datetime.strptime(raw_date, fmt).date()
                if parsed.year < 100:
                    parsed = parsed.replace(year=2000 + parsed.year)
                return parsed
            except ValueError:
                pass
        m = _DATE_RE.fullmatch(raw_date)
        if m:
            d = int(m.group(1))
            mm = int(m.group(2))
            yy = int(m.group(3)) if m.group(3) else (year or datetime.utcnow().year)
            if yy < 100:
                yy += 2000
            try:
                return date(yy, mm, d)
            except ValueError:
                return None
    if month and year and raw_date and raw_date.isdigit():
        try:
            return date(year, month, int(raw_date))
        except ValueError:
            return None
    return None


def _parse_hour(raw: Optional[str]) -> Optional[Tuple[int, int]]:
    if not raw:
        return None
    normalized_raw = _comparable_text(raw)
    m = _TIME_RE.search(normalized_raw)
    if not m:
        return None
    h = int(m.group(1))
    mm = int(m.group(2) or 0)
    if m.group(2) is None:
        compact = re.search(r"\b(\d{3,4})\b", normalized_raw)
        if compact:
            digits = compact.group(1)
            if len(digits) == 3:
                h, mm = int(digits[0]), int(digits[1:])
            else:
                h, mm = int(digits[:2]), int(digits[2:])
    if h > 23 or mm > 59:
        return None
    return h, mm


def _classify_shift_kind(start: Optional[Tuple[int, int]], end: Optional[Tuple[int, int]], shift_label: Optional[str]) -> str:
    label = _comparable_text(shift_label)
    if "24" in label and "hora" in label:
        return "twenty_four"
    if start == (8, 0) and end == (20, 0):
        return "day"
    if start == (10, 0) and end == (22, 0):
        return "intermediate"
    if start == (20, 0) and end in {(8, 0), (7, 0)}:
        return "night"
    return "custom"


def _parse_time_window(start_raw: Optional[str], end_raw: Optional[str], shift_label: Optional[str]) -> Tuple[Optional[Tuple[int, int]], Optional[Tuple[int, int]], str]:
    if start_raw and end_raw:
        start = _parse_hour(start_raw)
        end = _parse_hour(end_raw)
        return start, end, _classify_shift_kind(start, end, shift_label)

    label = _normalize_text(shift_label)
    m = re.search(r"(\d{1,2}(?::\d{2}|h\d{0,2})?)\s*(?:-|às|as|a)\s*(\d{1,2}(?::\d{2}|h\d{0,2})?)", label, re.IGNORECASE)
    if m:
        start = _parse_hour(m.group(1))
        end = _parse_hour(m.group(2))
        return start, end, _classify_shift_kind(start, end, shift_label)
    cmp = _comparable_text(label)
    if "24 horas" in cmp:
        return (0, 0), (0, 0), "twenty_four"
    if "noite" in cmp or "noturno" in cmp:
        return (20, 0), (8, 0), "night"
    if "manha" in cmp:
        return (8, 0), (14, 0), "day"
    if "tarde" in cmp:
        return (14, 0), (20, 0), "custom"
    if "dia" in cmp:
        return (8, 0), (20, 0), "day"
    return None, None, "custom"


def _detect_layout(headers: List[str], rows: List[List[Any]]) -> str:
    cmp_headers = {_comparable_text(h) for h in headers}
    has_avive = {"cidade", "estado", "empresa", "unidade", "especialidade", "profissional", "data", "dia"}.issubset(cmp_headers)
    has_pa = {"data", "dia", "plantao", "crm", "nome completo"}.issubset(cmp_headers)
    if has_pa:
        return "pa24h_block"
    if has_avive:
        return "avive_tabular"

    zeros = 0
    times = 0
    for row in rows[:20]:
        for cell in row:
            val = _normalize_text(str(cell) if cell is not None else "")
            if val == "00:00":
                zeros += 1
            if re.fullmatch(r"\d{1,2}:\d{2}", val):
                times += 1
    if zeros >= 6 and times >= 10:
        return "avive_tabular"
    if any("plant" in h for h in cmp_headers) and any("crm" in h for h in cmp_headers):
        return "pa24h_block"
    return "generic_table"


def _score_name_similarity(name_a: str, name_b: str) -> float:
    key = (_comparable_text(name_a), _comparable_text(name_b))
    if key in _MATCH_SCORE_CACHE:
        return _MATCH_SCORE_CACHE[key]
    score = SequenceMatcher(None, key[0], key[1]).ratio()
    _MATCH_SCORE_CACHE[key] = score
    return score


def _build_pattern_index(db: Session) -> Dict[int, Dict[str, int]]:
    result: Dict[int, Dict[str, int]] = {}
    recent = db.query(Shift).order_by(Shift.start_time.desc()).limit(400).all()
    for shift in recent:
        if not shift.agent_id:
            continue
        day = shift.start_time.weekday()
        hour = shift.start_time.hour
        key = f"{day}:{hour}"
        bucket = result.setdefault(int(shift.agent_id), {})
        bucket[key] = bucket.get(key, 0) + 1
    return result


def _resolve_user_match(db: Session, name: str, canonical_name: str, crm: Optional[str], shift_kind: str, parsed_date: Optional[date], context: Optional[Dict[str, Any]] = None) -> Tuple[str, Optional[int], List[str], float, Optional[str], Dict[str, Any], Optional[Dict[str, Any]]]:
    notes: List[str] = []
    context = context or {}
    users: List[User] = context.get("users") or db.query(User).filter(User.is_active == True).all()  # noqa: E712
    profiles: List[MedicalProfile] = context.get("profiles") or db.query(MedicalProfile).all()
    pattern_index: Dict[int, Dict[str, int]] = context.get("pattern_index") or _build_pattern_index(db)

    if not name:
        return "invalid", None, ["Nome obrigatório"], 0.0, None, {}, None
    if len(_normalize_text(name).split()) < 2:
        notes.append("Nome incompleto para match automático")
        return "new_user_candidate", None, notes, 0.35, None, {"name": 0.0, "alias": 0.0, "crm": 0.0, "pattern": 0.0, "shift": 0.0}, None

    normalized = _comparable_text(canonical_name or name)
    profile_by_crm = {str(p.crm_numero): p for p in profiles if p.crm_numero}

    score_components = {"name": 0.0, "alias": 0.0, "crm": 0.0, "pattern": 0.0, "shift": 0.0}
    suggested_profile_enrichment = None

    if crm:
        crm_digits = re.sub(r"\D", "", crm)
        profile = profile_by_crm.get(crm_digits)
        if profile:
            score_components["crm"] = 1.0
            notes.append("CRM exato encontrado")
            return "matched", profile.user_id, notes, 0.99, profile.nome_completo, score_components, None

    best_user = None
    best_score = 0.0
    for user in users:
        user_name = _comparable_text(user.name)
        score = _score_name_similarity(normalized, user_name)
        if score > best_score:
            best_user = user
            best_score = score

    if best_user:
        score_components["name"] = best_score
        if normalized != _comparable_text(name):
            score_components["alias"] = 0.1

        if parsed_date:
            pattern_key = f"{parsed_date.weekday()}:{8 if shift_kind == 'day' else 10 if shift_kind == 'intermediate' else 20 if shift_kind == 'night' else 0}"
            if pattern_index.get(best_user.id, {}).get(pattern_key, 0) > 1:
                score_components["pattern"] = 0.12
                notes.append("Coerente com pré-escala/recorrência histórica")

        if shift_kind in {"day", "intermediate", "night", "twenty_four"}:
            score_components["shift"] = 0.06

        profile = next((p for p in profiles if p.user_id == best_user.id), None)
        if crm and profile and re.sub(r"\D", "", crm) != str(profile.crm_numero):
            notes.append("Conflito de CRM, requer revisão")
            return "ambiguous", best_user.id, notes, 0.89, best_user.name, score_components, {"type": "crm_conflict", "crm_detected": crm}
        if crm and not profile:
            suggested_profile_enrichment = {"type": "add_crm", "user_id": best_user.id, "crm": crm}
            notes.append("Sugerido enriquecimento cadastral de CRM")

        if best_score >= 0.95:
            return "matched", best_user.id, notes, min(1.0, best_score + 0.1), best_user.name, score_components, suggested_profile_enrichment
        if best_score >= 0.85:
            notes.append("Match médio: revisão recomendada")
            return "ambiguous", best_user.id, notes, best_score, best_user.name, score_components, suggested_profile_enrichment

    if best_user and best_score < 0.85:
        notes.append("Baixa confiança de match")

    return "new_user_candidate", None, notes, max(0.3, best_score), None, score_components, suggested_profile_enrichment


def _build_row_from_values(
    db: Session,
    values: Dict[str, Any],
    source_sheet: Optional[str],
    source_page: Optional[int],
    source_table_index: Optional[int],
    row_index: int,
    month_ctx: Optional[int],
    year_ctx: Optional[int],
    source_layout_type: str,
    day_group_id: Optional[str],
    context: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    name_raw = values.get("professional")
    cleaned_name, crm_from_name, messages, canonical_name, alias_applied = _clean_professional(name_raw)
    crm_raw = values.get("crm") or crm_from_name

    parsed_date = _parse_date_with_context(values.get("date"), month_ctx, year_ctx)
    start_h, end_h, shift_kind = _parse_time_window(values.get("start_time"), values.get("end_time"), values.get("shift_label"))

    start_dt = end_dt = None
    duration = None
    if parsed_date and start_h and end_h:
        start_dt = datetime(parsed_date.year, parsed_date.month, parsed_date.day, start_h[0], start_h[1])
        end_dt = datetime(parsed_date.year, parsed_date.month, parsed_date.day, end_h[0], end_h[1])
        if shift_kind in {"night", "twenty_four"} or end_dt <= start_dt:
            end_dt += timedelta(days=1)
        duration = round((end_dt - start_dt).total_seconds() / 3600, 2)

    match_status, matched_user_id, match_notes, match_score, matched_name, score_components, enrichment = _resolve_user_match(
        db,
        cleaned_name,
        canonical_name,
        crm_raw,
        shift_kind,
        parsed_date,
        context=context,
    )
    messages.extend(match_notes)

    if not parsed_date:
        messages.append("Data ausente ou inválida")
    if not (start_h and end_h):
        messages.append("Horário/turno incompreensível")

    if not cleaned_name or not parsed_date or not (start_h and end_h):
        match_status = "invalid"

    confidence = 0.35
    if cleaned_name:
        confidence += 0.2
    if parsed_date:
        confidence += 0.2
    if start_h and end_h:
        confidence += 0.15
    if match_status == "matched":
        confidence += 0.1
    confidence = max(0.0, min(1.0, (confidence + match_score) / 2))

    return {
        "source_sheet": source_sheet,
        "source_page": source_page,
        "source_table_index": source_table_index,
        "source_row_index": row_index,
        "source_layout_type": source_layout_type,
        "day_group_id": day_group_id,
        "competency_month": parsed_date.month if parsed_date else month_ctx,
        "competency_year": parsed_date.year if parsed_date else year_ctx,
        "professional_name_raw": name_raw,
        "professional_name_normalized": cleaned_name,
        "canonical_name": canonical_name,
        "alias_applied": alias_applied,
        "crm_raw": values.get("crm"),
        "crm_normalized": crm_raw,
        "crm_detected": crm_raw,
        "crm_confidence": 0.95 if crm_raw else 0.0,
        "specialty_raw": values.get("specialty"),
        "location_raw": values.get("location"),
        "unit_raw": values.get("unit"),
        "date_raw": values.get("date"),
        "date_iso": parsed_date.isoformat() if parsed_date else None,
        "weekday_raw": values.get("weekday"),
        "shift_label_raw": values.get("shift_label"),
        "shift_kind": shift_kind,
        "schedule_pattern_type": "fixed" if score_components.get("pattern", 0) > 0 else "ad_hoc",
        "start_time_raw": values.get("start_time"),
        "end_time_raw": values.get("end_time"),
        "start_datetime": start_dt.isoformat() if start_dt else None,
        "end_datetime": end_dt.isoformat() if end_dt else None,
        "duration_hours": duration,
        "role_raw": values.get("role"),
        "notes": values.get("notes"),
        "confidence": round(confidence, 2),
        "match_status": match_status,
        "matched_user_id": matched_user_id,
        "suggested_existing_user_id": matched_user_id,
        "suggested_profile_enrichment": enrichment,
        "multiple_professionals_detected": bool(values.get("multiple_professionals_detected")),
        "grouped_day_validation": values.get("grouped_day_validation") or [],
        "match_score_components": score_components,
        "matched_name": matched_name,
        "validation_messages": list(dict.fromkeys(messages)),
        "llm_fallback_recommended": match_status in {"ambiguous", "invalid"} and confidence < 0.85,
    }


def _build_context(db: Session) -> Dict[str, Any]:
    return {
        "users": db.query(User).filter(User.is_active == True).all(),  # noqa: E712
        "profiles": db.query(MedicalProfile).all(),
        "pattern_index": _build_pattern_index(db),
    }


def _apply_avive_rules(headers: List[str], raw_row: List[Any]) -> Dict[str, Any]:
    header_lookup = {i: _comparable_text(h) for i, h in enumerate(headers)}
    values: Dict[str, Any] = {}
    raw_times: List[str] = []

    for col_idx, cell in enumerate(raw_row):
        cval = _normalize_text(str(cell) if cell is not None else "")
        header = header_lookup.get(col_idx, "")
        canonical = _canonical_header(header)
        if canonical:
            values[canonical] = cval
        if re.fullmatch(r"\d{1,2}:\d{2}", cval):
            raw_times.append(cval)

    if len(raw_times) >= 4 and raw_times[0] == "00:00" and raw_times[1] == "00:00":
        values["start_time"] = raw_times[-2]
        values["end_time"] = raw_times[-1]
    elif len(raw_times) >= 2:
        values["start_time"] = raw_times[-2]
        values["end_time"] = raw_times[-1]

    return values


def _build_rows_for_pa24h(raw_row: List[Any], canonical_by_idx: Dict[int, Optional[str]]) -> List[Dict[str, Any]]:
    values: Dict[str, Any] = {}
    for col_idx, cell in enumerate(raw_row):
        canonical = canonical_by_idx.get(col_idx)
        if canonical:
            values[canonical] = _normalize_text(str(cell))

    names = _split_compound_values(values.get("professional")) or [values.get("professional") or ""]
    crm_values = _split_compound_values(values.get("crm"))
    crm_candidates = []
    crm_raw = values.get("crm") or ""
    if crm_raw:
        crm_match = _CRM_NUMERIC_RE.findall(crm_raw)
        for g1, g2 in crm_match:
            if g1:
                crm_candidates.append(g1)
            if g2:
                crm_candidates.append(g2)
    if crm_values:
        crm_candidates = [re.sub(r"\D", "", crm) for crm in crm_values if crm]

    rows: List[Dict[str, Any]] = []
    for idx, name in enumerate(names):
        row_values = dict(values)
        row_values["professional"] = name
        row_values["multiple_professionals_detected"] = len(names) > 1
        if idx < len(crm_candidates):
            row_values["crm"] = crm_candidates[idx]
        elif crm_candidates:
            row_values["crm"] = crm_candidates[0]
            row_values.setdefault("grouped_day_validation", []).append("CRM composto sem alinhamento exato")
        rows.append(row_values)
    return rows


def _validate_day_groups(rows: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        day_id = row.get("day_group_id") or row.get("date_iso") or "unknown"
        grouped.setdefault(day_id, []).append(row)

    validation: Dict[str, List[str]] = {}
    for day_id, day_rows in grouped.items():
        messages: List[str] = []
        by_shift: Dict[str, int] = {}
        for row in day_rows:
            sk = row.get("shift_kind") or "custom"
            by_shift[sk] = by_shift.get(sk, 0) + 1
            if row.get("match_status") == "ambiguous":
                messages.append("Nome ambíguo")
            if row.get("crm_detected") is None:
                messages.append("CRM ausente")

        if any(count > 1 for shift, count in by_shift.items() if shift in {"day", "intermediate", "night"}):
            messages.append("Turno duplicado")
        if len(day_rows) > 6:
            messages.append("Número incomum de plantonistas")
        validation[day_id] = list(dict.fromkeys(messages))

    return validation


def normalize_xlsx_document(db: Session, content: bytes, filename: str) -> Dict[str, Any]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    rows: List[Dict[str, Any]] = []
    detected_months: List[Dict[str, Any]] = []
    raw_headers: List[str] = []
    normalized_headers: List[str] = []
    warnings: List[str] = []
    context = _build_context(db)

    for ws in wb.worksheets:
        matrix = list(ws.iter_rows(values_only=True))
        if not matrix:
            warnings.append(f"Aba '{ws.title}' vazia")
            continue

        best_header_idx = None
        best_score = -1
        best_header: List[str] = []
        for idx, row in enumerate(matrix[:15]):
            cells = [_normalize_text(str(c)) if c is not None else "" for c in row]
            score = sum(1 for cell in cells if _canonical_header(cell))
            if score > best_score:
                best_score = score
                best_header_idx = idx
                best_header = cells
        if best_header_idx is None or best_score < 2:
            warnings.append(f"Aba '{ws.title}' sem tabela válida detectada")
            continue

        layout_type = _detect_layout(best_header, [list(r) for r in matrix[best_header_idx + 1: best_header_idx + 16]])
        sheet_label = " ".join([str(c) for r in matrix[:best_header_idx + 1] for c in r if c])
        month, year, m_conf = _extract_month_year(f"{ws.title} {sheet_label}")
        if month and year:
            detected_months.append({"month": month, "year": year, "label": f"{month:02d}/{year}", "source_sheet": ws.title, "confidence": m_conf})

        header_map: Dict[int, str] = {}
        for i, h in enumerate(best_header):
            if h:
                raw_headers.append(h)
            canonical = _canonical_header(h)
            if canonical:
                header_map[i] = canonical
                normalized_headers.append(canonical)

        for offset, row in enumerate(matrix[best_header_idx + 1 :], start=best_header_idx + 2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            if layout_type == "avive_tabular":
                values = _apply_avive_rules(best_header, list(row))
            else:
                values = {}
                for col_idx, canonical in header_map.items():
                    if col_idx < len(row):
                        val = row[col_idx]
                        values[canonical] = _normalize_text(str(val)) if val is not None else None

            day_group_id = f"{values.get('date') or offset}"
            nrow = _build_row_from_values(db, values, ws.title, None, None, offset, month, year, layout_type, day_group_id, context=context)
            rows.append(nrow)

    wb.close()
    day_validation = _validate_day_groups(rows)
    for row in rows:
        row["grouped_day_validation"] = day_validation.get(row.get("day_group_id") or "", [])

    return {
        "source_type": "xlsx",
        "source_filename": filename,
        "detected_months": detected_months,
        "raw_headers": sorted(list(dict.fromkeys(raw_headers))),
        "normalized_headers": sorted(list(dict.fromkeys(normalized_headers))),
        "sheets": [{"name": ws.title} for ws in wb.worksheets],
        "rows": rows,
        "warnings": warnings,
        "errors": [],
        "metadata": {
            "detected_layout_type": "multi_sheet_tabular",
            "parser_version": "v3-deterministic-day-grouping",
            "ocr_provider": None,
            "overall_confidence": round(sum(r["confidence"] for r in rows) / len(rows), 2) if rows else 0.0,
            "llm_fallback_policy": "only_for_ambiguous_low_confidence",
        },
    }


def normalize_ocr_payload_document(db: Session, payload: Dict[str, Any], source_filename: str, source_type: str = "pdf") -> Dict[str, Any]:
    rows: List[Dict[str, Any]] = []
    warnings: List[str] = []
    detected_months: List[Dict[str, Any]] = []
    raw_headers: List[str] = []
    normalized_headers: List[str] = []
    context = _build_context(db)

    pages = payload.get("pages") or []
    for page in pages:
        page_number = page.get("page_number")
        for table_index, table in enumerate(page.get("tables") or [], start=1):
            title = _normalize_text(table.get("title"))
            month, year, conf = _extract_month_year(title)
            if month and year:
                detected_months.append({"month": month, "year": year, "label": title, "source_sheet": f"page-{page_number}", "confidence": conf})

            headers = [str(h) for h in (table.get("headers") or [])]
            raw_headers.extend(headers)
            canonical_by_idx = {i: _canonical_header(h) for i, h in enumerate(headers)}
            normalized_headers.extend([v for v in canonical_by_idx.values() if v])
            layout_type = _detect_layout(headers, table.get("rows") or [])

            for idx, raw_row in enumerate(table.get("rows") or [], start=1):
                candidate_rows: List[Dict[str, Any]] = []
                if layout_type == "avive_tabular":
                    values = _apply_avive_rules(headers, raw_row)
                    candidate_rows = [values]
                elif layout_type == "pa24h_block":
                    candidate_rows = _build_rows_for_pa24h(raw_row, canonical_by_idx)
                else:
                    values = {}
                    for col_idx, cell in enumerate(raw_row):
                        canonical = canonical_by_idx.get(col_idx)
                        if canonical:
                            values[canonical] = _normalize_text(str(cell))
                    candidate_rows = [values]

                for cidx, values in enumerate(candidate_rows):
                    parsed_date = _parse_date_with_context(values.get("date"), month, year)
                    day_group_id = parsed_date.isoformat() if parsed_date else f"page-{page_number}-row-{idx}"
                    source_row_index = idx if len(candidate_rows) == 1 else (idx * 100 + cidx)
                    row = _build_row_from_values(
                        db,
                        values,
                        f"page-{page_number}",
                        int(page_number) if page_number else None,
                        table_index,
                        source_row_index,
                        month,
                        year,
                        layout_type,
                        day_group_id,
                        context=context,
                    )
                    row_conf = table.get("confidence")
                    if isinstance(row_conf, (int, float)):
                        row["confidence"] = round((row["confidence"] + float(row_conf)) / 2, 2)
                    rows.append(row)

    if not pages:
        warnings.append("Payload OCR sem páginas")

    day_validation = _validate_day_groups(rows)
    for row in rows:
        row["grouped_day_validation"] = day_validation.get(row.get("day_group_id") or "", [])

    ambiguous_rows = [r for r in rows if r.get("llm_fallback_recommended")]

    return {
        "source_type": source_type,
        "source_filename": source_filename,
        "detected_months": detected_months,
        "raw_headers": sorted(list(dict.fromkeys(raw_headers))),
        "normalized_headers": sorted(list(dict.fromkeys(normalized_headers))),
        "sheets": [],
        "rows": rows,
        "warnings": warnings,
        "errors": [],
        "metadata": {
            "detected_layout_type": "ocr_table",
            "parser_version": "v3-deterministic-day-grouping",
            "ocr_provider": payload.get("provider") or "mockable",
            "overall_confidence": round(sum(r["confidence"] for r in rows) / len(rows), 2) if rows else 0.0,
            "llm_fallback_policy": "compact_ambiguous_rows_only",
            "llm_payload_candidates": len(ambiguous_rows),
            "shift_kind_colors": SHIFT_KIND_COLORS,
            "layout_counts": {
                "avive_tabular": sum(1 for r in rows if r.get("source_layout_type") == "avive_tabular"),
                "pa24h_block": sum(1 for r in rows if r.get("source_layout_type") == "pa24h_block"),
                "generic_table": sum(1 for r in rows if r.get("source_layout_type") == "generic_table"),
            },
        },
    }
