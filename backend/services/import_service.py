"""Serviço de importação de escala base.

Separa claramente:
  1. leitura do arquivo (CSV / Excel)
  2. parsing tabular
  3. normalização de turnos
  4. validação de consistência
  5. detecção de duplicatas e sobreposições
  6. persistência no banco (staging)
  7. confirmação: staging → Shift
  8. resumo da importação
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re
from datetime import date, datetime, time, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple

import httpx
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..config.settings import settings
from ..models.models import (
    ImportStatus,
    RowStatus,
    ScheduleImport,
    ScheduleImportRow,
    Shift,
    User,
)
from ..models.ocr_import import OcrImport

from .schedule_validation_service import validate_schedule, validate_shift

logger = logging.getLogger("agentescala.import_service")

# ─── Turnos padrão reconhecidos ──────────────────────────────────────────────
# (hora_inicio, hora_fim_no_mesmo_dia_ou_proximo)
STANDARD_SHIFTS: List[Tuple[time, time]] = [
    (time(8, 0),  time(20, 0)),
    (time(20, 0), time(8, 0)),   # vira dia
    (time(10, 0), time(22, 0)),
]

# ─── Aliases de colunas aceitos ──────────────────────────────────────────────
_COL_ALIASES: Dict[str, List[str]] = {
    "profissional":  ["profissional", "professional", "nome", "name", "agente", "agent", "colaborador"],
    "user_id":       ["user_id", "usuario_id", "id_usuario", "id_user", "profissional_id"],
    "data":          ["data", "date", "data_turno", "shift_date", "dia_data"],
    "dia_semana":    ["dia_semana", "dia", "day", "weekday", "semana"],
    "hora_inicio":   ["hora_inicio", "hora_inicial", "start_time", "inicio", "start", "start_hour",
                      "hora de inicio", "hora de início", "início"],
    "hora_fim":      ["hora_fim", "hora_final", "end_time", "fim", "end", "end_hour",
                      "hora de fim", "hora de termino", "hora de término", "término"],
    "total_horas":   ["total_horas", "horas", "hours", "duration", "duracao", "duração", "total"],
    "observacoes":   ["observacoes", "observação", "observacoes", "obs", "notes", "nota",
                      "observations", "observação"],
    "origem":        ["origem", "source", "origin", "fonte"],
}


def _build_column_map(headers: List[Any]) -> Dict[str, str]:
    """Mapeia cabeçalhos reais do arquivo para nomes canônicos."""
    norm = {str(h).strip().lower() if h else "": str(h) for h in headers}
    result: Dict[str, str] = {}
    for canonical, aliases in _COL_ALIASES.items():
        for alias in aliases:
            if alias in norm:
                result[canonical] = norm[alias]
                break
    return result


# ─── Parsing de data ─────────────────────────────────────────────────────────

_DATE_FORMATS = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d.%m.%Y"]


def _parse_date(raw: Optional[str]) -> Optional[date]:
    if not raw:
        return None
    raw = str(raw).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    # Tenta conversão para int (número serial do Excel)
    try:
        serial = int(float(raw))
        # Excel date serial: 1 = 1900-01-01
        base = date(1899, 12, 30)
        return base + timedelta(days=serial)
    except (ValueError, TypeError):
        pass
    return None


# ─── Parsing de hora ─────────────────────────────────────────────────────────

_TIME_FORMATS = ["%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"]


def _parse_time(raw: Optional[Any]) -> Optional[time]:
    if raw is None:
        return None
    # openpyxl pode devolver um objeto time diretamente
    if isinstance(raw, time):
        return raw
    if isinstance(raw, datetime):
        return raw.time()
    raw = str(raw).strip()
    if not raw:
        return None
    for fmt in _TIME_FORMATS:
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            pass
    return None


# ─── Leitura de arquivo ───────────────────────────────────────────────────────

def _read_csv(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    # Detecta separador
    sep = ";"
    first_line = text.split("\n")[0]
    if first_line.count(",") >= first_line.count(";"):
        sep = ","
    reader = csv.DictReader(io.StringIO(text), delimiter=sep)
    headers = reader.fieldnames or []
    rows = list(reader)
    return list(headers), rows


def _read_excel(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []
    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    if not headers_raw:
        return [], []
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(headers_raw)]
    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        rows.append(dict(zip(headers, row)))
    wb.close()
    return headers, rows


_DATE_PATTERN = re.compile(r"\b(?:\d{4}-\d{2}-\d{2}|\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b")
_TIME_PATTERN = re.compile(r"\b\d{1,2}:\d{2}(?::\d{2})?\b")
_OCR_SPLIT_PATTERN = re.compile(r"[|;,\t]| {2,}")


def _normalize_match_text(value: Optional[str]) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def _parse_ocr_line(raw_line: str) -> Tuple[Dict[str, Any], Optional[str]]:
    line = raw_line.strip()
    if not line:
        return {}, None

    date_match = _DATE_PATTERN.search(line)
    time_matches = _TIME_PATTERN.findall(line)

    parsed_date = date_match.group(0) if date_match else None
    parsed_start = time_matches[0] if len(time_matches) >= 1 else None
    parsed_end = time_matches[1] if len(time_matches) >= 2 else None

    parts = [p.strip() for p in _OCR_SPLIT_PATTERN.split(line) if p and p.strip()]
    name = None
    for part in parts:
        if part == parsed_date:
            continue
        if part in time_matches:
            continue
        if any(ch.isalpha() for ch in part):
            name = part
            break

    row = {
        "profissional": name,
        "data": parsed_date,
        "hora_inicio": parsed_start,
        "hora_fim": parsed_end,
        "observacoes": f"OCR raw: {line}",
        "origem": "ocr",
    }
    parse_error = None
    if not name or not parsed_date or not parsed_start or not parsed_end:
        parse_error = f"Linha OCR ambígua: '{line}'"
    return row, parse_error


def _parse_ocr_text_to_rows(raw_text: str) -> Tuple[List[str], List[Dict[str, Any]], List[Dict[str, str]]]:
    headers = ["profissional", "data", "hora_inicio", "hora_fim", "observacoes", "origem"]
    rows: List[Dict[str, Any]] = []
    errors: List[Dict[str, str]] = []

    for idx, line in enumerate(raw_text.splitlines(), start=1):
        parsed, parse_error = _parse_ocr_line(line)
        if not parsed:
            continue
        rows.append(parsed)
        if parse_error:
            errors.append({"code": "OCR_PARSE_AMBIGUOUS", "message": parse_error, "line": str(idx)})

    return headers, rows, errors


def _read_pdf_ocr(content: bytes) -> Tuple[List[str], List[Dict[str, Any]], Dict[str, Any]]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - depende de ambiente
        raise ValueError("Leitura OCR de PDF indisponível: instale a dependência 'pypdf'.") from exc

    reader = PdfReader(io.BytesIO(content))
    page_texts: List[str] = []
    for page in reader.pages:
        page_texts.append(page.extract_text() or "")
    full_text = "\n".join(page_texts).strip()
    headers, rows, errors = _parse_ocr_text_to_rows(full_text)
    return headers, rows, {"raw_text": full_text, "errors": errors}


def _read_image_ocr(content: bytes) -> Tuple[List[str], List[Dict[str, Any]], Dict[str, Any]]:
    try:
        import pytesseract  # type: ignore
        from PIL import Image  # type: ignore
    except ImportError as exc:  # pragma: no cover - depende de ambiente com OCR nativo
        raise ValueError(
            "OCR de imagem indisponível neste ambiente (dependências Pillow/pytesseract ausentes)."
        ) from exc

    image = Image.open(io.BytesIO(content))
    raw_text = pytesseract.image_to_string(image, lang="por+eng")
    headers, rows, errors = _parse_ocr_text_to_rows(raw_text)
    return headers, rows, {"raw_text": raw_text, "errors": errors}


def _extract_text_from_ocr_payload(payload: Dict[str, Any]) -> str:
    if isinstance(payload.get("raw_text"), str):
        return payload["raw_text"]
    if isinstance(payload.get("text"), str):
        return payload["text"]
    if isinstance(payload.get("content"), str):
        return payload["content"]

    lines = payload.get("lines")
    if isinstance(lines, list):
        return "\n".join(str(line) for line in lines if line is not None)

    data = payload.get("data")
    if isinstance(data, dict):
        return _extract_text_from_ocr_payload(data)

    result = payload.get("result")
    if isinstance(result, dict):
        return _extract_text_from_ocr_payload(result)

    return ""


def _read_ocr_via_api(content: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]], Dict[str, Any]]:
    base_url = (settings.OCR_API_BASE_URL or "").rstrip("/")
    if not settings.OCR_API_ENABLED or not base_url:
        raise ValueError("OCR API desabilitada por configuração.")

    endpoint_candidates = ("/ocr/extract", "/api/ocr/extract", "/extract")
    timeout = settings.OCR_API_TIMEOUT_SECONDS
    last_error: Optional[Exception] = None

    for endpoint in endpoint_candidates:
        target_url = f"{base_url}{endpoint}"
        try:
            with httpx.Client(timeout=timeout, verify=settings.OCR_API_VERIFY_SSL) as client:
                response = client.post(
                    target_url,
                    files={"file": (filename, content, "application/octet-stream")},
                )
            response.raise_for_status()
            payload = response.json()
            raw_text = _extract_text_from_ocr_payload(payload if isinstance(payload, dict) else {})
            if not raw_text.strip():
                raise ValueError(
                    "OCR API retornou payload sem conteúdo textual reconhecível; acionando fallback local."
                )
            headers, rows, errors = _parse_ocr_text_to_rows(raw_text)
            return headers, rows, {"raw_text": raw_text, "errors": errors, "source": target_url}
        except Exception as exc:  # pragma: no cover - integração externa
            last_error = exc
            logger.warning("Falha OCR API em %s: %s", target_url, exc)

    raise ValueError(f"OCR API indisponível em {base_url}: {last_error}")


def read_file(content: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    fn_lower = filename.lower()
    if fn_lower.endswith(".xlsx") or fn_lower.endswith(".xls"):
        return _read_excel(content)
    return _read_csv(content)


# ─── Matching de profissional ─────────────────────────────────────────────────

def _find_agent(db: Session, name: Optional[str], preferred_user_id: Optional[str] = None) -> Tuple[Optional[User], List[str], str]:
    if preferred_user_id:
        try:
            user_id = int(str(preferred_user_id).strip())
            by_id = db.query(User).filter(User.id == user_id, User.is_active == True).first()  # noqa: E712
            if by_id:
                return by_id, [], "resolved_by_user_id"
            return None, [f"user_id '{preferred_user_id}' não encontrado ou inativo"], "unmatched_user_id"
        except (TypeError, ValueError):
            return None, [f"user_id '{preferred_user_id}' inválido"], "invalid_user_id"

    if not name:
        return None, ["Profissional não informado"], "missing_name"

    name_clean = _normalize_match_text(name)
    users: List[User] = db.query(User).filter(User.is_active == True).all()  # noqa: E712
    for user in users:
        if _normalize_match_text(user.name) == name_clean or _normalize_match_text(user.email) == name_clean:
            return user, [], "exact_name"

    name_tokens: Set[str] = {token for token in name_clean.split(" ") if len(token) >= 3}
    matches: List[User] = []
    for user in users:
        user_name = _normalize_match_text(user.name)
        if name_clean in user_name or user_name in name_clean:
            matches.append(user)
            continue
        if name_tokens:
            overlap = len(name_tokens.intersection(set(user_name.split(" "))))
            if overlap >= 2:
                matches.append(user)

    if len(matches) == 1:
        return matches[0], [f"Profissional '{name}' resolvido por correspondência aproximada com '{matches[0].name}'"], "fuzzy_name"
    if len(matches) > 1:
        names = ", ".join(user.name for user in matches)
        return None, [f"Profissional '{name}' ambíguo: múltiplas correspondências ({names})"], "ambiguous_name"
    return None, [f"Profissional '{name}' não encontrado no sistema"], "not_found"


# ─── Normalização de turno ────────────────────────────────────────────────────

def _normalize_shift(
    shift_date: date,
    t_start: time,
    t_end: time,
) -> Tuple[datetime, datetime, int, bool, bool]:
    """
    Retorna (normalized_start, normalized_end, duration_minutes, is_overnight, is_standard).
    Trata virada de dia quando t_end <= t_start.
    """
    dt_start = datetime.combine(shift_date, t_start)
    dt_end = datetime.combine(shift_date, t_end)
    is_overnight = False
    if t_end <= t_start:
        dt_end += timedelta(days=1)
        is_overnight = True
    duration = int((dt_end - dt_start).total_seconds() / 60)
    is_standard = (t_start, t_end) in STANDARD_SHIFTS
    return dt_start, dt_end, duration, is_overnight, is_standard


# ─── Validação de linha ───────────────────────────────────────────────────────

def _validate_row(
    row_num: int,
    raw: Dict[str, Any],
    col_map: Dict[str, str],
    db: Session,
) -> Dict[str, Any]:
    """
    Valida e normaliza uma linha do arquivo.
    Retorna dicionário com campos normalizados e lista de issues.
    row_status: VALID / WARNING / INVALID
    """
    issues: List[str] = []
    fatal = False

    def get_raw(canonical: str) -> Optional[str]:
        mapped_col = col_map.get(canonical)
        if not mapped_col:
            return None
        val = raw.get(mapped_col)
        return str(val).strip() if val is not None and str(val).strip() else None

    raw_professional = get_raw("profissional")
    raw_user_id = get_raw("user_id")
    raw_date_str = get_raw("data")
    raw_start = get_raw("hora_inicio")
    raw_end = get_raw("hora_fim")
    raw_hours = get_raw("total_horas")
    raw_obs = get_raw("observacoes")
    raw_src = get_raw("origem")

    # Profissional
    agent, agent_issues, match_status = _find_agent(db, raw_professional, preferred_user_id=raw_user_id)
    issues.extend(agent_issues)
    if agent is None and not agent_issues:
        issues.append("Profissional inválido")
    if agent is None and raw_professional:
        # Profissional não resolvido - é um alerta mas não necessariamente fatal
        # (pode ser importado manualmente depois da confirmação)
        issues.append("Linha não poderá ser confirmada sem profissional válido")

    # Data
    parsed_date = _parse_date(raw_date_str)
    parse_status = "ok"
    if parsed_date is None:
        issues.append(f"Data inválida ou ausente: '{raw_date_str}'")
        fatal = True
        parse_status = "invalid"

    # Horas
    t_start: Optional[time] = None
    t_end: Optional[time] = None

    # Para xlsx, o campo pode já ser um time object
    raw_start_raw = raw.get(col_map.get("hora_inicio", "")) if col_map.get("hora_inicio") else None
    raw_end_raw = raw.get(col_map.get("hora_fim", "")) if col_map.get("hora_fim") else None

    t_start = _parse_time(raw_start_raw) if raw_start_raw is not None else _parse_time(raw_start)
    t_end = _parse_time(raw_end_raw) if raw_end_raw is not None else _parse_time(raw_end)

    if t_start is None:
        issues.append(f"Hora inicial inválida ou ausente: '{raw_start}'")
        fatal = True
        parse_status = "invalid"
    if t_end is None:
        issues.append(f"Hora final inválida ou ausente: '{raw_end}'")
        fatal = True
        parse_status = "invalid"

    # Normalização
    normalized_start = normalized_end = None
    duration_minutes = None
    is_overnight = is_standard = False

    if parsed_date and t_start and t_end:
        normalized_start, normalized_end, duration_minutes, is_overnight, is_standard = _normalize_shift(
            parsed_date, t_start, t_end
        )
        # Valida duração declarada
        if raw_hours:
            try:
                declared_h = float(raw_hours.replace(",", ".").replace("h", "").strip())
                declared_min = int(declared_h * 60)
                if abs(declared_min - duration_minutes) > 10:
                    issues.append(
                        f"Duração declarada ({declared_h}h) diverge do calculado "
                        f"({duration_minutes // 60}h{duration_minutes % 60:02d}min)"
                    )
            except (ValueError, AttributeError):
                pass

        # Duração suspeita (< 30 min ou > 24h)
        if duration_minutes < 30:
            issues.append(f"Duração muito curta: {duration_minutes} min")
            fatal = True
        elif duration_minutes > 24 * 60:
            issues.append(f"Duração excede 24h: {duration_minutes} min — possível erro na virada de dia")
            fatal = True

    row_status = RowStatus.INVALID if fatal else (RowStatus.WARNING if issues else RowStatus.VALID)
    validation_status = (
        "invalid" if row_status == RowStatus.INVALID
        else "warning" if row_status == RowStatus.WARNING
        else "valid"
    )
    if parse_status == "ok" and row_status == RowStatus.WARNING:
        parse_status = "warning"

    confidence_score = 0.0
    if raw_professional:
        confidence_score += 0.2
    if parsed_date:
        confidence_score += 0.25
    if t_start:
        confidence_score += 0.2
    if t_end:
        confidence_score += 0.2
    if agent is not None:
        confidence_score += 0.15
    if match_status == "resolved_by_user_id":
        confidence_score += 0.05
    if match_status == "ambiguous_name":
        confidence_score -= 0.2
    if fatal:
        confidence_score = min(confidence_score, 0.35)
    confidence_score = max(0.0, min(1.0, round(confidence_score, 2)))

    return {
        "row_number": row_num,
        "raw_professional": raw_professional,
        "raw_date": raw_date_str,
        "raw_start_time": raw_start,
        "raw_end_time": raw_end,
        "raw_total_hours": raw_hours,
        "raw_observations": raw_obs,
        "raw_source": raw_src,
        "agent_id": agent.id if agent else None,
        "normalized_start": normalized_start,
        "normalized_end": normalized_end,
        "duration_minutes": duration_minutes,
        "is_overnight": is_overnight,
        "is_standard_shift": is_standard,
        "row_status": row_status,
        "confidence_score": confidence_score,
        "parse_status": parse_status,
        "match_status": match_status,
        "validation_status": validation_status,
        "issues": json.dumps(issues) if issues else None,
        "is_duplicate": False,
        "has_overlap": False,
    }


# ─── Detecção de duplicatas e sobreposições ────────────────────────────────────

def _detect_duplicates_and_overlaps(validated: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Marca duplicatas e sobreposições intra-lote (mesmos dados ou mesmo agente/período)."""
    seen: Dict[Tuple, int] = {}  # (agent_id, start, end) → row_number

    for row in validated:
        if row["row_status"] == RowStatus.INVALID:
            continue
        aid = row["agent_id"]
        s = row["normalized_start"]
        e = row["normalized_end"]
        if aid is None or s is None or e is None:
            continue

        key = (aid, s, e)
        if key in seen:
            row["is_duplicate"] = True
            row["validation_status"] = "warning"
            issues_list = json.loads(row["issues"]) if row["issues"] else []
            issues_list.append(f"Possível duplicata da linha {seen[key]}")
            row["issues"] = json.dumps(issues_list)
            if row["row_status"] == RowStatus.VALID:
                row["row_status"] = RowStatus.WARNING
        else:
            seen[key] = row["row_number"]

    # Sobreposição: mesmo agente, intervalos que se intersectam mas não são idênticos
    by_agent: Dict[int, List[Dict[str, Any]]] = {}
    for row in validated:
        if row["agent_id"] and row["normalized_start"] and row["normalized_end"]:
            by_agent.setdefault(row["agent_id"], []).append(row)

    for agent_rows in by_agent.values():
        for i, a in enumerate(agent_rows):
            for b in agent_rows[i + 1:]:
                if a["is_duplicate"] or b["is_duplicate"]:
                    continue
                # Sobreposição: a.start < b.end AND b.start < a.end
                if a["normalized_start"] < b["normalized_end"] and b["normalized_start"] < a["normalized_end"]:
                    for row in (a, b):
                        row["has_overlap"] = True
                        row["validation_status"] = "conflict"
                        issues_list = json.loads(row["issues"]) if row["issues"] else []
                        msg = "Sobreposição de turno detectada com outra linha do mesmo agente neste lote"
                        if msg not in issues_list:
                            issues_list.append(msg)
                            row["issues"] = json.dumps(issues_list)
                        if row["row_status"] == RowStatus.VALID:
                            row["row_status"] = RowStatus.WARNING

    return validated


def _append_issue(row: ScheduleImportRow, message: str) -> None:
    issues_list = json.loads(row.issues) if row.issues else []
    if message not in issues_list:
        issues_list.append(message)
        row.issues = json.dumps(issues_list)
    if row.row_status == RowStatus.VALID:
        row.row_status = RowStatus.WARNING
    if row.validation_status in {"pending", "valid"}:
        row.validation_status = "warning"


def _status_equals(raw_status: Any, expected: RowStatus) -> bool:
    if isinstance(raw_status, RowStatus):
        return raw_status == expected
    status_text = str(raw_status).lower()
    return status_text == expected.value or status_text.endswith(f".{expected.value}")


def _refresh_ocr_counters_for_import(db: Session, schedule_import: ScheduleImport) -> None:
    if not schedule_import.source_description or "ocr_import_id:" not in schedule_import.source_description:
        return

    marker = schedule_import.source_description.split("ocr_import_id:")[-1].split("]")[0].strip()
    ocr_import = db.query(OcrImport).filter(OcrImport.id == marker).first()
    if not ocr_import:
        return

    rows = db.query(ScheduleImportRow).filter(ScheduleImportRow.import_id == schedule_import.id).all()
    ocr_import.extracted_lines = len(rows)
    ocr_import.valid_lines = sum(
        1 for row in rows
        if _status_equals(row.row_status, RowStatus.VALID) and row.agent_id is not None
    )
    ocr_import.ambiguous_lines = sum(
        1 for row in rows if "ambiguous" in (row.match_status or "") or row.agent_id is None
    )
    ocr_import.conflict_lines = sum(1 for row in rows if row.has_overlap or row.is_duplicate)


def _apply_schedule_validation_to_staging(db: Session, import_id: int) -> None:
    rows = db.query(ScheduleImportRow).filter(
        ScheduleImportRow.import_id == import_id,
        ScheduleImportRow.row_status != RowStatus.INVALID,
    ).all()

    candidate_rows = [
        row for row in rows
        if row.agent_id is not None and row.normalized_start is not None and row.normalized_end is not None
    ]

    if not candidate_rows:
        return

    existing_by_agent: Dict[int, List[Dict[str, Any]]] = {}
    for row in candidate_rows:
        if row.agent_id in existing_by_agent:
            continue
        existing_shifts = db.query(Shift).filter(Shift.agent_id == row.agent_id).all()
        existing_by_agent[row.agent_id] = [
            {
                "id": shift.id,
                "agent_id": shift.agent_id,
                "start_time": shift.start_time,
                "end_time": shift.end_time,
            }
            for shift in existing_shifts
        ]

    payload: List[Dict[str, Any]] = []
    included_existing_agents: set[int] = set()
    row_id_map: Dict[str, ScheduleImportRow] = {}
    for row in candidate_rows:
        row_key = f"import_row_{row.id}"
        row_id_map[row_key] = row
        if row.agent_id not in included_existing_agents:
            payload.extend(existing_by_agent[row.agent_id])
            included_existing_agents.add(row.agent_id)
        payload.append(
            {
                "id": row_key,
                "agent_id": row.agent_id,
                "start_time": row.normalized_start,
                "end_time": row.normalized_end,
            }
        )

    errors = validate_schedule(payload)
    if not errors:
        return

    for error in errors:
        code = error.get("code")
        message = error.get("message", "Erro de validação de escala")
        shift_id = error.get("shift_id")
        other_shift_id = error.get("other_shift_id")

        if shift_id in row_id_map:
            row = row_id_map[shift_id]
            row.has_overlap = row.has_overlap or code == "OVERLAPPING_SHIFTS"
            if code == "OVERLAPPING_SHIFTS":
                row.validation_status = "conflict"
            _append_issue(row, f"[{code}] {message}")
        if other_shift_id in row_id_map:
            row = row_id_map[other_shift_id]
            row.has_overlap = row.has_overlap or code == "OVERLAPPING_SHIFTS"
            if code == "OVERLAPPING_SHIFTS":
                row.validation_status = "conflict"
            _append_issue(row, f"[{code}] {message}")

        if code in {"DAILY_HOURS_EXCEEDED", "WEEKLY_HOURS_EXCEEDED"}:
            agent_id = error.get("agent_id")
            for row in candidate_rows:
                if row.agent_id == agent_id:
                    _append_issue(row, f"[{code}] {message}")


# ─── Persistência ─────────────────────────────────────────────────────────────

def process_import_file(
    db: Session,
    file_content: bytes,
    filename: str,
    reference_period: Optional[str],
    source_description: Optional[str],
    imported_by_id: int,
) -> ScheduleImport:
    """Lê, parseia, normaliza, valida e persiste um arquivo de importação."""

    schedule_import = ScheduleImport(
        filename=filename,
        reference_period=reference_period,
        source_description=source_description,
        status=ImportStatus.PROCESSING,
        imported_by=imported_by_id,
    )
    db.add(schedule_import)
    db.flush()  # obtém ID antes de inserir as linhas

    fn_lower = filename.lower()
    is_ocr_upload = fn_lower.endswith(".pdf") or fn_lower.endswith((".png", ".jpg", ".jpeg", ".webp", ".tiff"))
    ocr_import: Optional[OcrImport] = None

    try:
        ocr_meta: Dict[str, Any] = {}
        if fn_lower.endswith(".pdf"):
            try:
                headers, raw_rows, ocr_meta = _read_ocr_via_api(file_content, filename)
            except Exception as api_exc:
                logger.warning(
                    "OCR API indisponível para '%s'; usando fallback calibrado local. Motivo: %s",
                    filename,
                    api_exc,
                )
                headers, raw_rows, ocr_meta = _read_pdf_ocr(file_content)
                ocr_meta["fallback"] = "local_pdf"
        elif fn_lower.endswith((".png", ".jpg", ".jpeg", ".webp", ".tiff")):
            try:
                headers, raw_rows, ocr_meta = _read_ocr_via_api(file_content, filename)
            except Exception as api_exc:
                logger.warning(
                    "OCR API indisponível para '%s'; usando fallback calibrado local. Motivo: %s",
                    filename,
                    api_exc,
                )
                headers, raw_rows, ocr_meta = _read_image_ocr(file_content)
                ocr_meta["fallback"] = "local_image"
        else:
            headers, raw_rows = read_file(file_content, filename)

        if is_ocr_upload:
            parse_errors_by_line = {err.get("line") for err in (ocr_meta.get("errors") or [])}
            ocr_import = OcrImport(
                status="draft",
                file_name=filename,
                file_type="pdf" if fn_lower.endswith(".pdf") else "image",
                source_origin=source_description,
                processing_strategy=(
                    "ks-sm-api-ocr"
                    if ocr_meta.get("source")
                    else ("pypdf-extract-text" if fn_lower.endswith(".pdf") else "pytesseract-ocr")
                ),
                raw_payload={"text": ocr_meta.get("raw_text", "")},
                parsed_rows=[
                    {
                        "row_id": str(idx),
                        "raw_text": row.get("observacoes", ""),
                        "parsed_name": row.get("profissional"),
                        "matched_user_id": None,
                        "match_score": None,
                        "candidates": [],
                        "start_time": row.get("hora_inicio"),
                        "end_time": row.get("hora_fim"),
                        "location": None,
                        "row_status": "warning" if str(idx) in parse_errors_by_line else "valid",
                        "confidence_score": 0.5 if str(idx) in parse_errors_by_line else 0.75,
                        "parse_status": "warning" if str(idx) in parse_errors_by_line else "ok",
                        "match_status": "pending",
                        "validation_status": "pending",
                    }
                    for idx, row in enumerate(raw_rows, start=1)
                ],
                errors=ocr_meta.get("errors", []),
                action_log=[],
                created_by=imported_by_id,
                extracted_lines=len(raw_rows),
            )
            db.add(ocr_import)
            db.flush()
            schedule_import.source_description = (
                (source_description or "").strip() + f" [ocr_import_id:{ocr_import.id}]"
            ).strip()
    except Exception as exc:
        logger.exception("Falha ao ler arquivo de importação: %s", filename)
        schedule_import.status = ImportStatus.FAILED
        if ocr_import:
            ocr_import.status = "discarded"
            ocr_import.errors = (ocr_import.errors or []) + [{"code": "OCR_READ_ERROR", "message": str(exc)}]
        db.commit()
        raise ValueError(f"Não foi possível ler o arquivo '{filename}': {exc}") from exc

    if not raw_rows:
        schedule_import.status = ImportStatus.COMPLETED
        db.commit()
        return schedule_import

    col_map = _build_column_map(headers)

    # Validar que as colunas mínimas estão presentes
    missing = [c for c in ("profissional", "data", "hora_inicio", "hora_fim") if c not in col_map]
    if missing:
        schedule_import.status = ImportStatus.FAILED
        db.commit()
        raise ValueError(
            f"Arquivo sem colunas obrigatórias: {missing}. "
            f"Colunas encontradas: {headers}"
        )

    # Parsear e validar cada linha
    validated: List[Dict[str, Any]] = []
    for row_num, raw in enumerate(raw_rows, start=2):  # linha 1 = cabeçalho
        # Pula linhas completamente vazias
        values = [v for v in raw.values() if v is not None and str(v).strip()]
        if not values:
            continue
        row_data = _validate_row(row_num, raw, col_map, db)
        row_data["import_id"] = schedule_import.id
        validated.append(row_data)

    validated = _detect_duplicates_and_overlaps(validated)

    # Persistir linhas
    for row_data in validated:
        db_row = ScheduleImportRow(**row_data)
        db.add(db_row)
    db.flush()

    # Atualizar contadores
    total = len(validated)
    valid = sum(1 for r in validated if r["row_status"] == RowStatus.VALID)
    warning = sum(1 for r in validated if r["row_status"] == RowStatus.WARNING)
    invalid = sum(1 for r in validated if r["row_status"] == RowStatus.INVALID)
    duplicate = sum(1 for r in validated if r["is_duplicate"])

    schedule_import.total_rows = total
    schedule_import.valid_rows = valid
    schedule_import.warning_rows = warning
    schedule_import.invalid_rows = invalid
    schedule_import.duplicate_rows = duplicate
    schedule_import.status = ImportStatus.COMPLETED
    _apply_schedule_validation_to_staging(db, schedule_import.id)
    rows_after_validation = db.query(ScheduleImportRow).filter(ScheduleImportRow.import_id == schedule_import.id).all()
    schedule_import.valid_rows = sum(1 for r in rows_after_validation if _status_equals(r.row_status, RowStatus.VALID))
    schedule_import.warning_rows = sum(1 for r in rows_after_validation if _status_equals(r.row_status, RowStatus.WARNING))
    schedule_import.invalid_rows = sum(1 for r in rows_after_validation if _status_equals(r.row_status, RowStatus.INVALID))
    if ocr_import:
        ocr_import.status = "draft"
    _refresh_ocr_counters_for_import(db, schedule_import)

    db.commit()
    db.refresh(schedule_import)
    return schedule_import


# ─── Confirmação: staging → Shift ─────────────────────────────────────────────

def confirm_import(
    db: Session,
    import_id: int,
    confirmed_by_id: int,
) -> Tuple[ScheduleImport, int]:
    """
    Converte linhas válidas/warning (não duplicadas) em Shifts reais.
    Retorna (import, shifts_created).
    """
    schedule_import = db.query(ScheduleImport).filter(ScheduleImport.id == import_id).first()
    if not schedule_import:
        raise ValueError(f"Importação {import_id} não encontrada")
    if schedule_import.confirmed_at is not None:
        raise ValueError(f"Importação {import_id} já foi confirmada")

    importable_rows = db.query(ScheduleImportRow).filter(
        ScheduleImportRow.import_id == import_id,
        ScheduleImportRow.row_status.in_([RowStatus.VALID, RowStatus.WARNING]),
        ScheduleImportRow.is_duplicate == False,  # noqa: E712
        ScheduleImportRow.agent_id != None,  # noqa: E711
        ScheduleImportRow.normalized_start != None,  # noqa: E711
        ScheduleImportRow.normalized_end != None,  # noqa: E711
        ScheduleImportRow.created_shift_id == None,  # noqa: E711
    ).all()

    shifts_created = 0
    shifts_buffer: Dict[int, List[Shift]] = {}

    for row in importable_rows:
        agent_existing_shifts = shifts_buffer.get(row.agent_id)
        if agent_existing_shifts is None:
            agent_existing_shifts = db.query(Shift).filter(Shift.agent_id == row.agent_id).all()
            shifts_buffer[row.agent_id] = agent_existing_shifts

        validation_errors = validate_shift(
            {
                "agent_id": row.agent_id,
                "start_time": row.normalized_start,
                "end_time": row.normalized_end,
            },
            existing_shifts=agent_existing_shifts,
        )
        if validation_errors:
            row.has_overlap = True
            row.validation_status = "conflict"
            issues_list = json.loads(row.issues) if row.issues else []
            for error in validation_errors:
                issues_list.append(error.get("message", "Erro de validação de escala"))
            row.issues = json.dumps(list(dict.fromkeys(issues_list)))
            if row.row_status == RowStatus.VALID:
                row.row_status = RowStatus.WARNING
            continue

        obs_parts = []
        if row.raw_observations:
            obs_parts.append(row.raw_observations)
        obs_parts.append(f"[Importado de: {schedule_import.filename}]")

        shift = Shift(
            agent_id=row.agent_id,
            user_id=row.agent_id,
            legacy_agent_name=row.raw_professional,
            start_time=row.normalized_start,
            end_time=row.normalized_end,
            title="Turno importado",
            description=" | ".join(obs_parts),
        )
        db.add(shift)
        db.flush()
        row.created_shift_id = shift.id
        shifts_buffer.setdefault(row.agent_id, []).append(shift)
        shifts_created += 1

    schedule_import.confirmed_at = datetime.utcnow()
    schedule_import.confirmed_by = confirmed_by_id

    if schedule_import.source_description and "ocr_import_id:" in schedule_import.source_description:
        marker = schedule_import.source_description.split("ocr_import_id:")[-1].split("]")[0].strip()
        ocr_import = db.query(OcrImport).filter(OcrImport.id == marker).first()
        if ocr_import:
            ocr_import.status = "confirmed"
            ocr_import.confirmed_by = confirmed_by_id
            ocr_import.confirmed_at = datetime.utcnow()
    _refresh_ocr_counters_for_import(db, schedule_import)

    db.commit()
    db.refresh(schedule_import)
    return schedule_import, shifts_created


def validate_import_staging(db: Session, import_id: int) -> ScheduleImport:
    schedule_import = db.query(ScheduleImport).filter(ScheduleImport.id == import_id).first()
    if not schedule_import:
        raise ValueError(f"Importação {import_id} não encontrada")

    _apply_schedule_validation_to_staging(db, import_id)
    rows = db.query(ScheduleImportRow).filter(ScheduleImportRow.import_id == import_id).all()
    schedule_import.valid_rows = sum(1 for r in rows if _status_equals(r.row_status, RowStatus.VALID))
    schedule_import.warning_rows = sum(1 for r in rows if _status_equals(r.row_status, RowStatus.WARNING))
    schedule_import.invalid_rows = sum(1 for r in rows if _status_equals(r.row_status, RowStatus.INVALID))
    schedule_import.duplicate_rows = sum(1 for r in rows if r.is_duplicate)
    _refresh_ocr_counters_for_import(db, schedule_import)
    db.commit()
    db.refresh(schedule_import)
    return schedule_import


# ─── Relatório de inconsistências ─────────────────────────────────────────────

def export_issues_csv(db: Session, import_id: int) -> bytes:
    """Gera CSV com linhas que têm inconsistências."""
    rows = db.query(ScheduleImportRow).filter(
        ScheduleImportRow.import_id == import_id,
        ScheduleImportRow.row_status.in_([RowStatus.WARNING, RowStatus.INVALID]),
    ).order_by(ScheduleImportRow.row_number).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "linha", "profissional", "data", "hora_inicio", "hora_fim",
        "status", "duplicata", "sobreposicao", "problemas",
    ])
    for row in rows:
        issues_list = json.loads(row.issues) if row.issues else []
        writer.writerow([
            row.row_number,
            row.raw_professional or "",
            row.raw_date or "",
            row.raw_start_time or "",
            row.raw_end_time or "",
            row.row_status.value,
            "sim" if row.is_duplicate else "não",
            "sim" if row.has_overlap else "não",
            "; ".join(issues_list),
        ])
    return output.getvalue().encode("utf-8-sig")
