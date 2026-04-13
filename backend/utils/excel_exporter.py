from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List
from io import BytesIO
from datetime import datetime
from ..models import Shift, User


class ExcelExporter:
    """Exportador profissional de Excel para turnos"""

    @staticmethod
    def export_shifts(shifts: List[Shift], include_agent_info: bool = True) -> BytesIO:
        """
        Exportar turnos para Excel com formatação profissional

        Args:
            shifts: Lista de objetos Shift a serem exportados
            include_agent_info: Se deve incluir informações do agente

        Returns:
            BytesIO: Arquivo Excel em memória
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Turnos"

        # Define estilo de cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Define cabeçalhos
        if include_agent_info:
            headers = ["ID", "Nome do Agente", "E-mail do Agente", "Início", "Fim", "Duração (horas)", "Título", "Local", "Descrição"]
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 40
        else:
            headers = ["ID", "ID do Agente", "Início", "Fim", "Duração (horas)", "Título", "Local", "Descrição"]
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 40

        # Escreve cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Escreve dados
        for row_num, shift in enumerate(shifts, 2):
            duration = (shift.end_time - shift.start_time).total_seconds() / 3600

            if include_agent_info:
                data = [
                    shift.id,
                    shift.agent.name if shift.agent else "N/D",
                    shift.agent.email if shift.agent else "N/D",
                    shift.start_time.strftime("%Y-%m-%d %H:%M"),
                    shift.end_time.strftime("%Y-%m-%d %H:%M"),
                    round(duration, 2),
                    shift.title,
                    shift.location or "",
                    shift.description or ""
                ]
            else:
                data = [
                    shift.id,
                    shift.agent_id,
                    shift.start_time.strftime("%Y-%m-%d %H:%M"),
                    shift.end_time.strftime("%Y-%m-%d %H:%M"),
                    round(duration, 2),
                    shift.title,
                    shift.location or "",
                    shift.description or ""
                ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical="center")

        # Adiciona aba de metadados
        meta_ws = wb.create_sheet("Metadados")
        meta_ws['A1'] = "Data de exportação"
        meta_ws['B1'] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        meta_ws['A2'] = "Total de turnos"
        meta_ws['B2'] = len(shifts)
        meta_ws['A3'] = "Gerado por"
        meta_ws['B3'] = "Sistema AgentEscala"

        meta_ws['A1'].font = Font(bold=True)
        meta_ws['A2'].font = Font(bold=True)
        meta_ws['A3'].font = Font(bold=True)
        meta_ws.column_dimensions['A'].width = 20
        meta_ws.column_dimensions['B'].width = 30

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_swap_requests(swap_requests: List) -> BytesIO:
        """
        Exportar solicitações de troca para Excel

        Args:
            swap_requests: Lista de objetos SwapRequest a serem exportados

        Returns:
            BytesIO: Arquivo Excel em memória
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Solicitações de Troca"

        # Define estilo de cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Define cabeçalhos
        headers = [
            "ID", "Status", "Solicitante", "Agente Alvo",
            "Turno de Origem", "Turno de Destino", "Motivo",
            "Notas do Admin", "Criado em", "Atualizado em"
        ]

        # Define larguras de coluna
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20

        # Escreve cabeçalhos
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Escreve dados
        for row_num, swap in enumerate(swap_requests, 2):
            data = [
                swap.id,
                swap.status.value,
                swap.requester.name if swap.requester else "N/D",
                swap.target_agent.name if swap.target_agent else "N/D",
                f"Turno #{swap.origin_shift_id}",
                f"Turno #{swap.target_shift_id}",
                swap.reason or "",
                swap.admin_notes or "",
                swap.created_at.strftime("%Y-%m-%d %H:%M"),
                swap.updated_at.strftime("%Y-%m-%d %H:%M")
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical="center")

        # Adiciona aba de metadados
        meta_ws = wb.create_sheet("Metadados")
        meta_ws['A1'] = "Data de exportação"
        meta_ws['B1'] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        meta_ws['A2'] = "Total de solicitações"
        meta_ws['B2'] = len(swap_requests)
        meta_ws['A3'] = "Gerado por"
        meta_ws['B3'] = "Sistema AgentEscala"

        meta_ws['A1'].font = Font(bold=True)
        meta_ws['A2'].font = Font(bold=True)
        meta_ws['A3'].font = Font(bold=True)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
