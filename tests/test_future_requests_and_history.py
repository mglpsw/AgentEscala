from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook

from backend.config.database import SessionLocal
from backend.models import Shift


def test_shifts_list_supports_past_present_future_with_date_filters(client, agent_headers, admin_headers):
    # cria um turno no passado para garantir histórico consultável
    db = SessionLocal()
    try:
        alice_id = db.query(Shift).first().agent_id
        past_start = datetime.utcnow().replace(hour=8, minute=0, second=0, microsecond=0) - timedelta(days=35)
        past_shift = Shift(
            agent_id=alice_id,
            start_time=past_start,
            end_time=past_start + timedelta(hours=8),
            title='Plantão histórico',
            location='Histórico',
        )
        db.add(past_shift)
        db.commit()
    finally:
        db.close()

    today = date.today()
    start = today.replace(day=1)
    response = client.get('/shifts', headers=agent_headers, params={'start_date': start.isoformat()})
    assert response.status_code == 200
    assert all(item['start_time'][:10] >= start.isoformat() for item in response.json())

    past_only = client.get(
        '/shifts',
        headers=agent_headers,
        params={'start_date': (today - timedelta(days=50)).isoformat(), 'end_date': (today - timedelta(days=20)).isoformat()},
    )
    assert past_only.status_code == 200
    titles = [item['title'] for item in past_only.json()]
    assert 'Plantão histórico' in titles


def test_create_list_cancel_future_shift_request_own_user(client, agent_headers):
    target_date = (date.today() + timedelta(days=20)).isoformat()

    create_response = client.post(
        '/me/future-shift-requests',
        headers=agent_headers,
        json={
            'requested_date': target_date,
            'shift_period': '12H DIA',
            'notes': 'Preferência para planejamento',
        },
    )
    assert create_response.status_code == 201
    created = create_response.json()
    assert created['requested_date'] == target_date
    assert created['status'] == 'active'

    list_response = client.get('/me/future-shift-requests', headers=agent_headers)
    assert list_response.status_code == 200
    listed_ids = [item['id'] for item in list_response.json()]
    assert created['id'] in listed_ids

    cancel_response = client.delete(f"/me/future-shift-requests/{created['id']}", headers=agent_headers)
    assert cancel_response.status_code == 200
    assert cancel_response.json()['status'] == 'cancelled'


def test_future_shift_request_rejects_more_than_three_months_ahead(client, agent_headers):
    too_far = (date.today() + timedelta(days=100)).isoformat()
    response = client.post(
        '/me/future-shift-requests',
        headers=agent_headers,
        json={
            'requested_date': too_far,
            'shift_period': '12H DIA',
        },
    )
    assert response.status_code == 400
    assert '3 meses' in response.json()['detail']


def test_user_cannot_cancel_other_users_future_request(client, admin_headers, agent_headers):
    bob_login = client.post(
        '/auth/login',
        json={'email': 'bob@agentescala.com', 'password': 'CHANGE_ME_TEST_PASSWORD'},
    )
    bob_headers = {'Authorization': f"Bearer {bob_login.json()['access_token']}"}

    created = client.post(
        '/me/future-shift-requests',
        headers=bob_headers,
        json={
            'requested_date': (date.today() + timedelta(days=10)).isoformat(),
            'shift_period': '12H NOITE',
        },
    )
    assert created.status_code == 201
    request_id = created.json()['id']

    forbidden = client.delete(f'/me/future-shift-requests/{request_id}', headers=agent_headers)
    assert forbidden.status_code == 404


def test_monthly_consolidated_xlsx_export_contains_required_columns(client, agent_headers):
    today = date.today()
    response = client.get(
        '/shifts/export/monthly-consolidated',
        headers=agent_headers,
        params={'year': today.year, 'month': today.month},
    )
    assert response.status_code == 200
    assert 'spreadsheetml' in response.headers['content-type']

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook['Consolidado Mensal']
    headers = [sheet.cell(row=1, column=index).value for index in range(1, 8)]
    assert headers == [
        'Data',
        'Dia da semana',
        'Turno',
        'Profissional',
        'Categoria profissional',
        'Status',
        'Local',
    ]
